#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV Export Module for Penalty Tracking System
Exports penalty data from Excel workbook to CSV format
"""

import openpyxl
import csv
from datetime import datetime
import os


def export_penalties_to_csv(excel_filename="Strafenerfassung_ASV_Natz.xlsx", csv_filename="Erfassung_Export.csv"):
    """
    Export penalty data from Excel workbook to CSV file
    
    Args:
        excel_filename (str): Path to the Excel workbook
        csv_filename (str): Path for the output CSV file
    
    Returns:
        str: Path to the created CSV file
    """
    
    # Check if Excel file exists
    if not os.path.exists(excel_filename):
        raise FileNotFoundError(f"Excel file '{excel_filename}' not found!")
    
    try:
        # Load workbook
        print(f"📖 Lade Excel-Datei: {excel_filename}")
        workbook = openpyxl.load_workbook(excel_filename, data_only=True)
        
        # Get the Erfassung worksheet
        if "Erfassung" not in workbook.sheetnames:
            raise ValueError("Arbeitsblatt 'Erfassung' nicht gefunden!")
        
        worksheet = workbook["Erfassung"]
        print(f"✅ Arbeitsblatt 'Erfassung' gefunden")
        
        # Find the table boundaries
        # Headers should be in row 2
        headers = []
        max_col = 0
        
        # Get headers from row 2
        for col in range(1, 8):  # A to G
            cell_value = worksheet.cell(row=2, column=col).value
            if cell_value:
                headers.append(str(cell_value))
                max_col = col
            else:
                break
        
        if not headers:
            raise ValueError("Keine Kopfzeilen in Zeile 2 gefunden!")
        
        print(f"📋 Gefundene Spalten: {headers}")
        
        # Collect data rows (only filled rows)
        data_rows = []
        processed_rows = 0
        filled_rows = 0
        
        # Start from row 3 (data rows)
        for row in range(3, worksheet.max_row + 1):
            row_data = []
            has_data = False
            
            # Check each column for this row
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value
                
                # Format the value appropriately
                if cell_value is None:
                    formatted_value = ""
                elif isinstance(cell_value, datetime):
                    # Format dates as YYYY-MM-DD
                    formatted_value = cell_value.strftime("%Y-%m-%d")
                    has_data = True
                elif isinstance(cell_value, (int, float)):
                    # Format numbers without currency symbols, use decimal point
                    if cell_value == 0 and col > 3:  # Don't count zero amounts as data
                        formatted_value = "0"
                    else:
                        formatted_value = str(cell_value).replace(',', '.')
                        if col <= 3:  # Date, Player, Penalty columns count as data
                            has_data = True
                        elif cell_value > 0:  # Only positive amounts count as data
                            has_data = True
                else:
                    # String values
                    formatted_value = str(cell_value).strip()
                    if formatted_value:  # Non-empty strings count as data
                        has_data = True
                
                row_data.append(formatted_value)
            
            processed_rows += 1
            
            # Only add row if it contains actual data (not just formulas with 0 results)
            if has_data:
                # Additional check: at least date OR player OR penalty must be filled
                date_filled = row_data[0] if len(row_data) > 0 else ""
                player_filled = row_data[1] if len(row_data) > 1 else ""
                penalty_filled = row_data[2] if len(row_data) > 2 else ""
                
                if date_filled or player_filled or penalty_filled:
                    data_rows.append(row_data)
                    filled_rows += 1
        
        print(f"📊 Verarbeitete Zeilen: {processed_rows}")
        print(f"📋 Befüllte Zeilen gefunden: {filled_rows}")
        
        # Write to CSV file
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            # Use semicolon as delimiter (German standard)
            writer = csv.writer(csvfile, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            
            # Write headers
            writer.writerow(headers)
            
            # Write data rows
            for row_data in data_rows:
                writer.writerow(row_data)
        
        print(f"✅ CSV-Export erfolgreich: {csv_filename}")
        print(f"📁 Exportierte Datensätze: {len(data_rows)}")
        
        # Show sample of exported data
        if data_rows:
            print(f"\n📋 Beispiel der ersten exportierten Zeilen:")
            print(f"Kopfzeilen: {'; '.join(headers)}")
            for i, row in enumerate(data_rows[:3]):  # Show first 3 rows
                print(f"Zeile {i+1}: {'; '.join(row)}")
            if len(data_rows) > 3:
                print(f"... und {len(data_rows)-3} weitere Zeilen")
        else:
            print("⚠️  Keine Daten zum Exportieren gefunden")
        
        workbook.close()
        return csv_filename
        
    except Exception as e:
        print(f"❌ Fehler beim CSV-Export: {str(e)}")
        raise


def validate_csv_export(csv_filename="Erfassung_Export.csv"):
    """
    Validate the exported CSV file
    
    Args:
        csv_filename (str): Path to the CSV file to validate
        
    Returns:
        dict: Validation results
    """
    
    if not os.path.exists(csv_filename):
        return {"valid": False, "error": f"CSV-Datei '{csv_filename}' nicht gefunden"}
    
    try:
        with open(csv_filename, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=';')
            
            # Read headers
            headers = next(reader)
            row_count = 0
            
            # Count data rows
            for row in reader:
                row_count += 1
            
            return {
                "valid": True,
                "headers": headers,
                "row_count": row_count,
                "encoding": "utf-8",
                "delimiter": "semicolon"
            }
            
    except Exception as e:
        return {"valid": False, "error": str(e)}


if __name__ == "__main__":
    try:
        # Export penalties to CSV
        excel_file = "Strafenerfassung_ASV_Natz.xlsx"
        csv_file = "Erfassung_Export.csv"
        
        print("🚀 Starte CSV-Export...")
        result_file = export_penalties_to_csv(excel_file, csv_file)
        
        # Validate the export
        validation = validate_csv_export(result_file)
        
        if validation["valid"]:
            print(f"\n✅ CSV-Export Validierung erfolgreich:")
            print(f"   📋 Kopfzeilen: {len(validation['headers'])}")
            print(f"   📊 Datensätze: {validation['row_count']}")
            print(f"   🔤 Kodierung: {validation['encoding']}")
            print(f"   📄 Trennzeichen: {validation['delimiter']}")
        else:
            print(f"\n❌ CSV-Export Validierung fehlgeschlagen: {validation['error']}")
            
    except FileNotFoundError as e:
        print(f"❌ Datei nicht gefunden: {str(e)}")
        print("💡 Tipp: Führen Sie zuerst 'python build_strafenlog.py' aus, um die Excel-Datei zu erstellen.")
    except Exception as e:
        print(f"❌ Unerwarteter Fehler: {str(e)}")
