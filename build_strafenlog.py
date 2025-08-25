#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Penalty Tracking System for ASV Natz Football Team
Generates comprehensive Excel workbook with penalty tracking, statistics, and analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date
import os


def create_penalty_tracking_workbook():
    """Create the complete penalty tracking Excel workbook"""
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default worksheet
    wb.remove(wb.active)
    
    # Create all worksheets
    ws_erfassung = wb.create_sheet("Erfassung")
    ws_spielerliste = wb.create_sheet("Spielerliste")
    ws_strafenkatalog = wb.create_sheet("Strafenkatalog")
    ws_statistik = wb.create_sheet("Statistik")
    ws_trainingsplan = wb.create_sheet("Trainingsplan")
    
    # Set active sheet to Erfassung
    wb.active = ws_erfassung
    
    # Define styles
    create_styles(wb)
    
    # Create each sheet
    create_erfassung_sheet(ws_erfassung)
    create_spielerliste_sheet(ws_spielerliste)
    create_strafenkatalog_sheet(ws_strafenkatalog)
    create_statistik_sheet(ws_statistik)
    create_trainingsplan_sheet(ws_trainingsplan)
    
    # Save workbook
    filename = "Strafenerfassung_ASV_Natz.xlsx"
    wb.save(filename)
    print(f"✅ Excel-Datei '{filename}' erfolgreich erstellt!")
    
    return filename


def create_styles(wb):
    """Create named styles for consistent formatting"""
    
    # Header style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="000000")
    header_style.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_style.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(header_style)
    
    # Currency style
    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = '#,##0.00 [$€-de-DE]'
    wb.add_named_style(currency_style)


def create_erfassung_sheet(ws):
    """Create the main penalty input sheet"""
    
    # Info message in first row
    ws.merge_cells('H1:P1')
    ws['H1'] = "Datum, Spieler & Vergehen wählen – Rest füllt sich automatisch. Filter nutzen, um Zeitraum/Spieler zu filtern."
    ws['H1'].font = Font(italic=True, size=9)
    ws['H1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Datum", "Spieler", "Vergehen", "Anzahl", "Einzelbetrag (€)", "Gesamt (€)", "Notiz"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # Set column widths
    column_widths = [13, 24, 36, 10, 18, 16, 28]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add data rows (1500 rows from row 3 to 1502)
    for row in range(3, 1503):
        # Datum (A) - no default value
        ws.cell(row=row, column=1)
        
        # Spieler (B) - no default value
        ws.cell(row=row, column=2)
        
        # Vergehen (C) - no default value
        ws.cell(row=row, column=3)
        
        # Anzahl (D) - default value 1
        ws.cell(row=row, column=4, value=1)
        
        # Einzelbetrag (E) - formula
        formula_e = f'=IFERROR(XLOOKUP(C{row},Strafenkatalog!$A$2:$A$400,Strafenkatalog!$B$2:$B$400),IFERROR(VLOOKUP(C{row},Strafenkatalog!$A$2:$B$400,2,FALSE),0))'
        ws.cell(row=row, column=5, value=formula_e)
        ws.cell(row=row, column=5).style = "currency_style"
        
        # Gesamt (F) - formula
        formula_f = f'=IFERROR(D{row}*E{row},0)'
        ws.cell(row=row, column=6, value=formula_f)
        ws.cell(row=row, column=6).style = "currency_style"
        
        # Notiz (G) - empty
        ws.cell(row=row, column=7)
    
    # Create table
    table_range = "A2:G1502"
    table = Table(displayName="tblErfassung", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    
    # Data validation for Datum
    date_validation = DataValidation(type="date", operator="between", 
                                   formula1=date(2000, 1, 1), formula2=date(2100, 12, 31),
                                   errorTitle="Ungültiges Datum", 
                                   error="Bitte geben Sie ein gültiges Datum zwischen 01.01.2000 und 31.12.2100 ein.")
    date_validation.add(f"A3:A1502")
    ws.add_data_validation(date_validation)
    
    # Data validation for Spieler (dropdown)
    player_validation = DataValidation(type="list", formula1="=Spielerliste!$A$2:$A$200",
                                     errorTitle="Ungültiger Spieler",
                                     error="Bitte wählen Sie einen Spieler aus der Liste.")
    player_validation.add(f"B3:B1502")
    ws.add_data_validation(player_validation)
    
    # Data validation for Vergehen (dropdown)
    penalty_validation = DataValidation(type="list", formula1="=Strafenkatalog!$A$2:$A$400",
                                       errorTitle="Ungültiges Vergehen",
                                       error="Bitte wählen Sie ein Vergehen aus dem Katalog.")
    penalty_validation.add(f"C3:C1502")
    ws.add_data_validation(penalty_validation)
    
    # Data validation for Anzahl
    anzahl_validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=1,
                                      errorTitle="Ungültige Anzahl",
                                      error="Die Anzahl muss mindestens 1 betragen.")
    anzahl_validation.add(f"D3:D1502")
    ws.add_data_validation(anzahl_validation)
    
    # Conditional formatting
    # Yellow fill for empty dates
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    empty_date_rule = FormulaRule(formula=[f'AND(A3="",OR(B3<>"",C3<>""))'], fill=yellow_fill)
    ws.conditional_formatting.add(f"A3:A1502", empty_date_rule)
    
    # Red fill for empty player or penalty
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    empty_player_rule = FormulaRule(formula=[f'AND(B3="",OR(A3<>"",C3<>""))'], fill=red_fill)
    empty_penalty_rule = FormulaRule(formula=[f'AND(C3="",OR(A3<>"",B3<>""))'], fill=red_fill)
    ws.conditional_formatting.add(f"B3:B1502", empty_player_rule)
    ws.conditional_formatting.add(f"C3:C1502", empty_penalty_rule)
    
    # Green fill for positive amounts
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    positive_amount_rule = CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
    ws.conditional_formatting.add(f"F3:F1502", positive_amount_rule)
    
    # Enable filter
    ws.auto_filter.ref = table_range
    
    # Freeze panes
    ws.freeze_panes = 'A3'


def create_spielerliste_sheet(ws):
    """Create the player list sheet"""
    
    # Header
    ws['A1'] = "Spieler"
    ws['A1'].style = "header_style"
    
    # Player names
    players = [
        "Maximilian Hofer", "Hannes Peintner", "Alex Braunhofer", "Alex Schraffel",
        "Andreas Fusco", "Armin Feretti", "Hannes Larcher", "Julian Brunner",
        "Leo Tauber", "Lukas Mayr", "Manuel Troger", "Martin Gasser",
        "Matthias Schmid", "Maximilian Schraffl", "Michael Mitterrutzner", "Michael Peintner",
        "Patrick Auer", "Patrick Pietersteiner", "Stefan Filo", "Stefan Peintner",
        "Manuel Auer", "Mauro Monti", "Tobias", "Jakob Unterholzner",
        "Fabian Bacher", "Emil Gabrieli", "Mardochee", "Oleg Schleiermann"
    ]
    
    # Add players to sheet
    for i, player in enumerate(players, 2):
        ws.cell(row=i, column=1, value=player)
    
    # Set column width
    ws.column_dimensions['A'].width = 26
    
    # Create table
    table_range = "A1:A200"
    table = Table(displayName="tblSpieler", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def create_strafenkatalog_sheet(ws):
    """Create the penalty catalog sheet"""
    
    # Headers
    headers = ["Vergehen", "Strafe (€) pro Einheit", "Beschreibung (optional)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "header_style"
    
    # Penalty data
    penalties = [
        ("Unentschuldigtes Fehlen im Trainingslager", 50, ""),
        ("Bier bei Essen Trainingslager", 10, ""),
        ("Busfahrer pflanzen", 5, ""),
        ("Alpha Aktion", 5, ""),
        ("Ball in Q5", 2, ""),
        ("Socken ohschneiden", 20, ""),
        ("Valentinstog fahln", 50, ""),
        ("Abschlussmatch verloren", 2, ""),
        ("Fehlen beim Spiel wegen Urlaub", 30, ""),
        ("Abwesenheit Urlaub während Meisterschaft", 10, ""),
        ("Unentschuldigtes Fehlen Spiel", 50, ""),
        ("Unentschieden Meisterschaftsspiel", 1, ""),
        ("Niederlage Meisterschaftsspiel", 2, ""),
        ("Spiel Socken ohschneiden", 20, ""),
        ("Elfer verursachen", 10, ""),
        ("Unentschuldigtes Fehlen beim Training", 20, ""),
        ("100%ige Chance liegen lossen", 5, ""),
        ("Falscher Einwurf", 5, ""),
        ("Elfer verschiaßn", 10, ""),
        ("Tormonn Papelle kregn", 5, ""),
        ("Freitig glei nochn Training gian", 2, ""),
        ("Schuache in Kabine ohklopfn", 5, ""),
        ("Kistenplan net einholten /pro Kopf", 30, ""),
        ("Übung bei training vertschecken", 1, ""),
        ("Torello 20 Pässe", 2, ""),
        ("Übern tennisplotz mit FB schuach gian", 5, ""),
        ("Nochn training gian ohne eps zu verraumen", 5, ""),
        ("Gelbsperre/Rotsperre pro Spiel", 15, ""),
        ("Kabinendienst vernachlässigt", 10, ""),
        ("Freitags Abschluss-Spiel verloren", 2, ""),
        ("Abwesenheit Urlaub in Vorbereitung", 5, ""),
        ("Glei nochn Hoamspiel gian(min 30 min.)", 10, ""),
        ("Saufn vorn Spiel", 50, ""),
        ("Unsportliches Verhalten gegenüber Mitspieler/Trai", 50, ""),
        ("Erstes Tor/Startelfeinsatz", 0, "Kasten (ansonsten 20€)"),
        ("Eigentor", 0, "Kasten (ansonsten 20€)"),
        ("Foto in Zeitung/Online", 2, ""),
        ("Sachen in Kabine/Platz vergessen", 5, ""),
        ("Unentschuldigtes fehlen beim Training ohne Absage", 15, ""),
        ("Rauchen im Trikot", 15, ""),
        ("Bei Spiel folscher Trainer", 20, ""),
        ("Folsches Trainingsgewond", 5, ""),
        ("Handy leitn in do kabine", 5, ""),
        ("Schiffn in do Dusche", 20, ""),
        ("Oan setzn in do Kabine (wenns stinkt 20€)", 5, ""),
        ("Frau/freindin fa an Mitspieler verraumen", 500, ""),
        ("Geburtstogsessen net innerholb 1 Monat gebrocht", 150, ""),
        ("Rote Karte wegn Unsportlichkeit", 50, ""),
        ("Gelbe Karte wegn Unsportlichkeit", 20, ""),
        ("Zu spät - Pauschale", 5, "")
    ]
    
    # Add penalty data
    for i, (penalty, amount, description) in enumerate(penalties, 2):
        ws.cell(row=i, column=1, value=penalty)
        ws.cell(row=i, column=2, value=amount).style = "currency_style"
        ws.cell(row=i, column=3, value=description)
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    
    # Create table with 200 additional empty rows for future entries
    table_range = "A1:C400"
    table = Table(displayName="tblKatalog", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def create_statistik_sheet(ws):
    """Create the comprehensive statistics sheet"""
    
    # Control section headers
    ws['A2'] = "Zeitraum Start"
    ws['A2'].style = "header_style"
    ws['B2'] = f"=DATE(YEAR(TODAY()),MONTH(TODAY()),1)"
    
    ws['C2'] = "Zeitraum Ende"
    ws['C2'].style = "header_style"
    ws['D2'] = "=TODAY()"
    
    ws['A4'] = "Spieler (Auswahl)"
    ws['A4'].style = "header_style"
    
    # Player selection dropdown
    player_validation = DataValidation(type="list", formula1="=Spielerliste!$A$2:$A$200")
    player_validation.add("B4")
    ws.add_data_validation(player_validation)
    
    # Set column widths for control section
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    
    # KPIs section
    ws['A6'] = "Gesamtbetrag (Zeitraum)"
    ws['A6'].style = "header_style"
    ws['B6'] = '=IFERROR(SUMIFS(tblErfassung[Gesamt (€)],tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0)'
    ws['B6'].style = "currency_style"
    
    ws['A7'] = "Anzahl Einträge (Zeitraum)"
    ws['A7'].style = "header_style"
    ws['B7'] = '=IFERROR(COUNTIFS(tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0)'
    
    ws['A8'] = "Ø Betrag pro Eintrag"
    ws['A8'].style = "header_style"
    ws['B8'] = '=IF(B7=0,0,B6/B7)'
    ws['B8'].style = "currency_style"
    
    ws['A9'] = "Höchste Einzelstrafe"
    ws['A9'].style = "header_style"
    ws['B9'] = '=IFERROR(AGGREGATE(14,6,tblErfassung[Gesamt (€)]/(tblErfassung[Datum]>=$B$2)/(tblErfassung[Datum]<=$D$2),1),0)'
    ws['B9'].style = "currency_style"
    
    # Player statistics table
    ws['A11'] = "Spieler"
    ws['A11'].style = "header_style"
    ws['B11'] = "Summe (€)"
    ws['B11'].style = "header_style"
    ws['C11'] = "Anzahl"
    ws['C11'].style = "header_style"
    ws['D11'] = "Ø (€)"
    ws['D11'].style = "header_style"
    
    # Add players and formulas
    players = [
        "Maximilian Hofer", "Hannes Peintner", "Alex Braunhofer", "Alex Schraffel",
        "Andreas Fusco", "Armin Feretti", "Hannes Larcher", "Julian Brunner",
        "Leo Tauber", "Lukas Mayr", "Manuel Troger", "Martin Gasser",
        "Matthias Schmid", "Maximilian Schraffl", "Michael Mitterrutzner", "Michael Peintner",
        "Patrick Auer", "Patrick Pietersteiner", "Stefan Filo", "Stefan Peintner",
        "Manuel Auer", "Mauro Monti", "Tobias", "Jakob Unterholzner",
        "Fabian Bacher", "Emil Gabrieli", "Mardochee", "Oleg Schleiermann"
    ]
    
    for i, player in enumerate(players, 12):
        ws.cell(row=i, column=1, value=player)
        ws.cell(row=i, column=2, value=f'=IF(A{i}="",0,IFERROR(SUMIFS(tblErfassung[Gesamt (€)],tblErfassung[Spieler],A{i},tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0))').style = "currency_style"
        ws.cell(row=i, column=3, value=f'=IF(A{i}="",0,IFERROR(COUNTIFS(tblErfassung[Spieler],A{i},tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0))')
        ws.cell(row=i, column=4, value=f'=IF(C{i}=0,0,B{i}/C{i})').style = "currency_style"
    
    # Penalty statistics table
    ws['F11'] = "Vergehen"
    ws['F11'].style = "header_style"
    ws['G11'] = "Summe (€)"
    ws['G11'].style = "header_style"
    ws['H11'] = "Anzahl"
    ws['H11'].style = "header_style"
    ws['I11'] = "Ø (€)"
    ws['I11'].style = "header_style"
    
    # Add penalty formulas (using first 50 penalties as example)
    penalties = [
        "Unentschuldigtes Fehlen im Trainingslager", "Bier bei Essen Trainingslager", "Busfahrer pflanzen",
        "Alpha Aktion", "Ball in Q5", "Socken ohschneiden", "Valentinstog fahln", "Abschlussmatch verloren",
        "Fehlen beim Spiel wegen Urlaub", "Abwesenheit Urlaub während Meisterschaft"
    ]
    
    for i, penalty in enumerate(penalties[:10], 12):  # Show first 10 as example
        ws.cell(row=i, column=6, value=penalty)
        ws.cell(row=i, column=7, value=f'=IF(F{i}="",0,IFERROR(SUMIFS(tblErfassung[Gesamt (€)],tblErfassung[Vergehen],F{i},tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0))').style = "currency_style"
        ws.cell(row=i, column=8, value=f'=IF(F{i}="",0,IFERROR(COUNTIFS(tblErfassung[Vergehen],F{i},tblErfassung[Datum],">="&$B$2,tblErfassung[Datum],"<="&$D$2),0))')
        ws.cell(row=i, column=9, value=f'=IF(H{i}=0,0,G{i}/H{i})').style = "currency_style"
    
    # Create player statistics table
    table_range = "A11:D40"
    table = Table(displayName="tblStatSpieler", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    
    # Create penalty statistics table
    table_range = "F11:I21"
    table = Table(displayName="tblStatVergehen", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    
    # Time series section
    ws['A25'] = "Strafen über Zeit – Spieler"
    ws['A25'].font = Font(bold=True)
    ws['F24'] = "Tage anzeigen"
    ws['F24'].style = "header_style"
    ws['G24'] = 90
    
    # Date series and amounts for time series
    for i in range(26, 116):  # 90 days
        ws.cell(row=i, column=1, value=f'=IF(ROW()-ROW($A$26)+1<=$G$24,$B$2+ROW()-ROW($A$26),"")')
        ws.cell(row=i, column=2, value=f'=IF(A{i}="","",IFERROR(SUMIFS(tblErfassung[Gesamt (€)],tblErfassung[Spieler],$B$4,tblErfassung[Datum],A{i}),0))')
    
    # Create time series chart
    chart = LineChart()
    chart.title = "Strafen über Zeit – Spieler"
    chart.y_axis.title = "Betrag (€)"
    chart.x_axis.title = "Datum"
    
    # Chart data
    dates = Reference(ws, min_col=1, min_row=26, max_row=115)
    amounts = Reference(ws, min_col=2, min_row=26, max_row=115)
    chart.add_data(amounts, titles_from_data=False)
    chart.set_categories(dates)
    
    # Position and size the chart
    chart.width = 26 * 7  # Approximately 26 columns width
    chart.height = 12 * 15  # Approximately 12 rows height
    ws.add_chart(chart, "A36")
    
    # Monthly matrix section
    ws['A65'] = "Monatsmatrix (Summe €)"
    ws['A65'].font = Font(bold=True)
    
    # Month headers
    for col in range(2, 14):  # 12 months
        col_letter = get_column_letter(col)
        if col == 2:
            ws[f'{col_letter}65'] = '=EOMONTH($B$2,0)'
        else:
            prev_col = get_column_letter(col-1)
            ws[f'{col_letter}65'] = f'=EOMONTH({prev_col}65,1)'
        ws[f'{col_letter}65'].number_format = 'MMM YYYY'
    
    # Player names and monthly data
    for i, player in enumerate(players, 66):
        ws.cell(row=i, column=1, value=player)
        for col in range(2, 14):  # 12 months
            col_letter = get_column_letter(col)
            formula = f'=IF($A{i}="",0,IFERROR(SUMIFS(tblErfassung[Gesamt (€)],tblErfassung[Spieler],$A{i},tblErfassung[Datum],">="&EOMONTH({col_letter}$65,-1)+1,tblErfassung[Datum],"<="&EOMONTH({col_letter}$65,0)),0))'
            ws.cell(row=i, column=col, value=formula).style = "currency_style"
    
    # Freeze panes
    ws.freeze_panes = 'A12'


def create_trainingsplan_sheet(ws):
    """Create the training schedule sheet"""
    
    # Headers
    headers = ["Datum", "Tag", "Uhrzeit", "Einheit", "Schuhe"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "header_style"
    
    # Sample training data
    training_data = [
        ("2024-08-01", "Donnerstag", "19:00", "Konditionstraining", "Stollen"),
        ("2024-08-03", "Samstag", "10:00", "Testspiel", "Stollen"),
        ("2024-08-05", "Montag", "19:30", "Techniktraining", "Halle"),
        ("2024-08-08", "Donnerstag", "19:00", "Spielaufbau", "Stollen"),
        ("2024-08-10", "Samstag", "15:00", "Meisterschaftsspiel", "Stollen")
    ]
    
    # Add training data
    for i, (datum, tag, zeit, einheit, schuhe) in enumerate(training_data, 2):
        ws.cell(row=i, column=1, value=datum)
        ws.cell(row=i, column=2, value=tag)
        ws.cell(row=i, column=3, value=zeit)
        ws.cell(row=i, column=4, value=einheit)
        ws.cell(row=i, column=5, value=schuhe)
    
    # Set column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Create table
    table_range = "A1:E20"
    table = Table(displayName="tblTrainingsplan", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


if __name__ == "__main__":
    try:
        filename = create_penalty_tracking_workbook()
        print(f"Penalty tracking system successfully created: {filename}")
    except Exception as e:
        print(f"Error creating workbook: {str(e)}")
        raise
